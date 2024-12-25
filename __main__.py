import argparse
import shutil
from datetime import datetime

import openpyxl
import vdf

from model import Achievement, Game
from utils import steam, xlsx
from pathlib import Path

parser = argparse.ArgumentParser()

parser.add_argument("--oldresult", default=None, help="Use previous result for concatenation", action="store")
parser.add_argument("--output", default="./output.xlsx", help="Dir or file to store", action="store")

args = parser.parse_args()

schema_path = Path(f"~/.steam/root/appcache/stats/").expanduser().resolve().absolute()
workbook = None

old_path = None
new_path = None

if args.oldresult is not None:
    old_path = Path(args.oldresult).expanduser().resolve().absolute()
    if not old_path.exists():
        print("old result not found")
        exit(1)
    if not old_path.is_file():
        print("old result should be a file")
        exit(1)

if args.output is not None:
    new_path = Path(args.output).expanduser().resolve().absolute()

    if not new_path.parent.exists():
        print("output directory not found")
        exit(1)

if old_path == new_path:
    old_path_new = old_path.parent / (old_path.stem + f"_{datetime.now().isoformat()}" + ".".join(old_path.suffixes))
    old_path = Path(shutil.move(old_path, old_path_new)).resolve().absolute()

if old_path:
    workbook = openpyxl.load_workbook(old_path)

games_database = steam.get_apps_info()
if len(games_database) == 0:
    print("Can't get games")
    exit(1)

games_appid_cache = set()
games_priced_dict = {}
games_free_dict = {}
games_removed_dict = {}

processed_games_database = {}

if workbook is not None:
    for sheet_name, games_dict in zip(["Priced", "Free", "Removed"],
                                      [games_priced_dict, games_free_dict, games_removed_dict]):
        if sheet_name not in workbook.sheetnames:
            continue

        for sheet_row in workbook[sheet_name].iter_rows(min_row=2, values_only=True):
            appid = str(sheet_row[0])
            name = str(sheet_row[1])
            protected_count = int(sheet_row[2])
            achievement_count = int(sheet_row[3])
            if sheet_row[4] is not None and protected_count > 0:
                achievement_list = sheet_row[4].split(";")
            else:
                achievement_list = []

            processed_games_database[appid] = Game(
                name=name,
                appid=appid,
                has_protected=protected_count > 0,
                achievement_count=achievement_count,
                protected_count=protected_count,
                achievements={name: Achievement(protected=True, name=name) for name in achievement_list}
            )
            games_dict[appid] = processed_games_database[appid]

request_count = 0
for file in schema_path.glob("UserGameStatsSchema_*.bin"):
    achievements_out = dict()
    stats_out = dict()
    with open(file, 'rb') as f:
        schema = vdf.binary_loads(f.read())

    appid = None
    appname = None
    has_protected = False
    protected_count = 0

    for app_id, app_stats in schema.items():
        if not appid:
            appid = app_id
        if not appname:
            appname = app_stats.get("gamename", app_stats.get("GameName", None))

        if not isinstance(app_stats['stats'], dict):
            continue

        for stat in app_stats['stats'].values():
            if stat['type'] == '4':
                for achievement in stat['bits'].values():
                    new_achievement = steam.parse_achievement_schema(achievement)
                    has_protected = has_protected or new_achievement.protected
                    if new_achievement.protected:
                        protected_count += 1
                    achievements_out[achievement['name']] = new_achievement

    if appid in processed_games_database:
        continue

    if not appname:
        appname = ""

    if appid not in games_database:
        new_game = Game(name=appname, appid=appid)
        processed_games_database[appid] = new_game
    else:
        processed_games_database[appid] = games_database[appid]
        if not processed_games_database[appid]:
            processed_games_database[appid] = processed_games_database[appid]

    processed_games_database[appid].achievements = achievements_out
    processed_games_database[appid].has_protected = has_protected
    processed_games_database[appid].protected_count = protected_count
    processed_games_database[appid].achievement_count = len(achievements_out)

    request_count += 1
    appdetails = steam.request_price(appid)
    if not appdetails:
        continue

    if "success" in appdetails and appdetails["success"]:
        if "data" in appdetails:
            if "name" in appdetails["data"]:
                processed_games_database[appid].name = appdetails["data"]["name"]

            if "price_overview" in appdetails["data"]:
                games_priced_dict[appid] = processed_games_database[appid]
            elif "is_free" in appdetails["data"] and not appdetails["data"]["is_free"]:
                games_removed_dict[appid] = processed_games_database[appid]
            else:
                games_free_dict[appid] = processed_games_database[appid]
        else:
            games_free_dict[appid] = processed_games_database[appid]
    else:
        games_removed_dict[appid] = processed_games_database[appid]


def data_sort(game):
    return not game.has_protected, game.name


games_priced_list = sorted(games_priced_dict.values(), key=data_sort)
games_free_list = sorted(games_free_dict.values(), key=data_sort)
games_removed_list = sorted(games_removed_dict.values(), key=data_sort)

if not workbook:
    workbook = openpyxl.Workbook()

for sheet_name in ["Priced", "Free", "Removed"]:
    if sheet_name not in workbook.sheetnames:
        workbook.create_sheet(sheet_name)

for worksheet_name, games_list in zip(["Priced", "Free", "Removed"],
                                      [games_priced_list, games_free_list, games_removed_list]):
    write_count = xlsx.sheet_writer(workbook[worksheet_name], games_list)
    workbook[worksheet_name].delete_rows(write_count + 1, workbook[worksheet_name].max_row)

workbook.save(new_path)
